from __future__ import annotations

import base64
import csv
import io
import json
import re
import sys
import zipfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass
class ExtractionResult:
    text: str
    parse_mode: str
    source_format: str
    warnings: list[dict[str, str]]
    metadata: dict[str, Any]


class _HtmlTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self.list_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_map = {key.lower(): value for key, value in attrs}
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            level = int(tag[1])
            self.parts.append("\n" + ("#" * level) + " ")
        elif tag == "li":
            self.parts.append("\n- ")
        elif tag == "blockquote":
            self.parts.append("\n> ")
        elif tag == "br":
            self.parts.append("\n")
        elif tag == "img":
            alt = (attrs_map.get("alt") or "").strip()
            src = (attrs_map.get("src") or "").strip()
            self.parts.append(f"\n![{alt}]({src})\n")
        elif tag in {"p", "div", "section", "article", "table", "tr"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"p", "div", "section", "article", "table", "tr", "blockquote"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if data:
            self.parts.append(data)

    def text(self) -> str:
        normalized = "".join(self.parts)
        normalized = re.sub(r"\n{3,}", "\n\n", normalized)
        return html_unescape(normalized).strip()


def main() -> None:
    payload = json.loads(sys.stdin.read() or "{}")
    result = extract(payload)
    json.dump(
        {
            "text": result.text,
            "parseMode": result.parse_mode,
            "sourceFormat": result.source_format,
            "warnings": result.warnings,
            "metadata": result.metadata,
        },
        sys.stdout,
    )


def extract(payload: dict[str, Any]) -> ExtractionResult:
    title = str(payload.get("title") or "document")
    mime_kind = str(payload.get("mimeKind") or "text/plain")
    source_format = str(payload.get("sourceFormat") or infer_format(title, mime_kind))
    encoding = str(payload.get("encoding") or "text")
    warnings: list[dict[str, str]] = []

    try:
      raw_bytes = (
          base64.b64decode(str(payload.get("content") or ""), validate=False)
          if encoding == "base64"
          else str(payload.get("content") or "").encode("utf8")
      )
    except Exception as exc:  # pragma: no cover - defensive path
      raise RuntimeError(f"Could not decode document payload: {exc}") from exc

    text_content = raw_bytes.decode("utf8", errors="replace") if encoding != "base64" else ""

    if source_format in {"plain_text", "txt", "markdown", "latex", "typst", "json", "yaml", "tsv"}:
        parse_mode = "markdown" if source_format in {"markdown", "typst", "latex"} else source_format
        return ExtractionResult(
            text=text_content,
            parse_mode=parse_mode,
            source_format=source_format,
            warnings=warnings,
            metadata={"contentEncoding": encoding},
        )

    if source_format == "html":
        return ExtractionResult(
            text=html_to_markdown_like(text_content),
            parse_mode="markdown",
            source_format=source_format,
            warnings=warnings,
            metadata={"contentEncoding": encoding},
        )

    if source_format == "csv":
        rows = parse_csv_rows(text_content, ",")
        return ExtractionResult(
            text=rows_to_pipe_table(title, rows),
            parse_mode="markdown",
            source_format=source_format,
            warnings=warnings,
            metadata={"rowCount": max(len(rows) - 1, 0), "contentEncoding": encoding},
        )

    if source_format == "pdf":
        return extract_pdf(raw_bytes, title, warnings, encoding)
    if source_format == "docx":
        return extract_docx(raw_bytes, title, warnings, encoding)
    if source_format == "xlsx":
        return extract_xlsx(raw_bytes, title, warnings, encoding)
    if source_format == "epub":
        return extract_epub(raw_bytes, title, warnings, encoding)

    warnings.append(
        {
            "code": "UNSUPPORTED_FORMAT_FALLBACK",
            "message": f"Falling back to plain-text handling for {source_format}.",
        }
    )
    return ExtractionResult(
        text=text_content,
        parse_mode="plain_text",
        source_format=source_format,
        warnings=warnings,
        metadata={"contentEncoding": encoding},
    )


def extract_pdf(
    raw_bytes: bytes, title: str, warnings: list[dict[str, str]], encoding: str
) -> ExtractionResult:
    reader = PdfReader(io.BytesIO(raw_bytes))
    parts = [f"# {title}"]
    extracted_pages = 0
    for index, page in enumerate(reader.pages):
        page_text = (page.extract_text() or "").strip()
        if not page_text:
            continue
        extracted_pages += 1
        parts.append(f"\n## Page {index + 1}\n\n{page_text}")
    if extracted_pages == 0:
        warnings.append(
            {
                "code": "OCR_REQUIRED",
                "message": "PDF text extraction returned no readable text. OCR may be required.",
            }
        )
    return ExtractionResult(
        text="\n".join(parts).strip(),
        parse_mode="markdown",
        source_format="pdf",
        warnings=warnings,
        metadata={
            "pageCount": len(reader.pages),
            "extractedPageCount": extracted_pages,
            "ocrStatus": "required" if extracted_pages == 0 else "not_needed",
            "contentEncoding": encoding,
        },
    )


def extract_docx(
    raw_bytes: bytes, title: str, warnings: list[dict[str, str]], encoding: str
) -> ExtractionResult:
    document = DocxDocument(io.BytesIO(raw_bytes))
    parts = [f"# {title}"]
    paragraphs = 0
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        paragraphs += 1
        if getattr(paragraph.style, "name", "").lower().startswith("heading"):
            level_match = re.search(r"(\d+)", getattr(paragraph.style, "name", ""))
            level = min(max(int(level_match.group(1)) if level_match else 1, 1), 6)
            parts.append("\n" + ("#" * level) + f" {text}")
        else:
            parts.append("\n" + text)
    if paragraphs == 0:
        warnings.append({"code": "EMPTY_DOCUMENT", "message": "DOCX contained no readable paragraphs."})
    return ExtractionResult(
        text="\n".join(parts).strip(),
        parse_mode="markdown",
        source_format="docx",
        warnings=warnings,
        metadata={"paragraphCount": paragraphs, "contentEncoding": encoding},
    )


def extract_xlsx(
    raw_bytes: bytes, title: str, warnings: list[dict[str, str]], encoding: str
) -> ExtractionResult:
    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    parts = [f"# {title}"]
    sheet_names: list[str] = []
    row_count = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        if not rows:
            continue
        sheet_names.append(sheet_name)
        row_count += max(len(rows) - 1, 0)
        parts.append(f"\n## Sheet: {sheet_name}\n")
        parts.append(rows_to_pipe_table(sheet_name, rows, include_title=False))
    if not sheet_names:
        warnings.append({"code": "EMPTY_WORKBOOK", "message": "Workbook contained no readable sheets."})
    return ExtractionResult(
        text="\n".join(part for part in parts if part.strip()).strip(),
        parse_mode="markdown",
        source_format="xlsx",
        warnings=warnings,
        metadata={"sheetNames": sheet_names, "rowCount": row_count, "contentEncoding": encoding},
    )


def extract_epub(
    raw_bytes: bytes, title: str, warnings: list[dict[str, str]], encoding: str
) -> ExtractionResult:
    parts = [f"# {title}"]
    chapter_count = 0
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
        html_names = [
            name
            for name in archive.namelist()
            if name.lower().endswith((".xhtml", ".html", ".htm"))
        ]
        for name in sorted(html_names):
            chapter_text = html_to_markdown_like(
                archive.read(name).decode("utf8", errors="replace")
            ).strip()
            if not chapter_text:
                continue
            chapter_count += 1
            parts.append(f"\n## {Path(name).stem}\n\n{chapter_text}")
    if chapter_count == 0:
        warnings.append({"code": "EMPTY_EPUB", "message": "EPUB contained no readable HTML chapters."})
    return ExtractionResult(
        text="\n".join(parts).strip(),
        parse_mode="markdown",
        source_format="epub",
        warnings=warnings,
        metadata={"chapterCount": chapter_count, "contentEncoding": encoding},
    )


def parse_csv_rows(text: str, delimiter: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[cell.strip() for cell in row] for row in reader if any(cell.strip() for cell in row)]


def rows_to_pipe_table(title: str, rows: list[list[str]], include_title: bool = True) -> str:
    if not rows:
        return f"# {title}" if include_title else ""
    normalized = [pad_row(row, max(len(item) for item in rows)) for row in rows]
    header = normalized[0]
    separator = ["---"] * len(header)
    body = normalized[1:]
    lines = []
    if include_title:
        lines.append(f"# {title}")
        lines.append("")
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(separator) + " |")
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines).strip()


def pad_row(row: list[str], width: int) -> list[str]:
    return row + [""] * max(width - len(row), 0)


def html_to_markdown_like(value: str) -> str:
    parser = _HtmlTextExtractor()
    parser.feed(value)
    return parser.text()


def infer_format(title: str, mime_kind: str) -> str:
    lowered_title = title.lower()
    if lowered_title.endswith((".md", ".markdown")):
        return "markdown"
    if lowered_title.endswith((".typ", ".typst")):
        return "typst"
    if lowered_title.endswith(".csv"):
        return "csv"
    if lowered_title.endswith(".tsv"):
        return "tsv"
    if lowered_title.endswith(".json"):
        return "json"
    if lowered_title.endswith((".yaml", ".yml")):
        return "yaml"
    if lowered_title.endswith(".html"):
        return "html"
    if lowered_title.endswith(".pdf"):
        return "pdf"
    if lowered_title.endswith(".docx"):
        return "docx"
    if lowered_title.endswith(".xlsx"):
        return "xlsx"
    if lowered_title.endswith(".epub"):
        return "epub"
    by_mime = {
        "text/markdown": "markdown",
        "text/html": "html",
        "text/csv": "csv",
        "text/tab-separated-values": "tsv",
        "application/pdf": "pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/epub+zip": "epub",
        "application/json": "json",
    }
    return by_mime.get(mime_kind, "plain_text")


def html_unescape(text: str) -> str:
    return (
        text.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )


if __name__ == "__main__":
    main()
